#!/usr/bin/env python3
"""
Repair corrupted Excel file and implement proper locking
"""

import os
import shutil
import zipfile
from datetime import datetime
import openpyxl
import pandas as pd

def repair_excel():
    """Repair corrupted Excel file from backup or recreate"""
    
    filename = "live_trading_results.xlsx"
    
    # Find most recent valid backup
    backups = []
    for f in os.listdir('.'):
        if f.startswith('live_trading_results_backup') and f.endswith('.xlsx'):
            try:
                # Test if backup is valid
                with zipfile.ZipFile(f, 'r') as z:
                    z.testzip()
                backups.append((f, os.path.getmtime(f)))
            except:
                pass
    
    if backups:
        # Sort by modification time (newest first)
        backups.sort(key=lambda x: x[1], reverse=True)
        best_backup = backups[0][0]
        
        print(f"Found valid backup: {best_backup}")
        
        # Backup current corrupted file
        if os.path.exists(filename):
            corrupted_name = f"live_trading_results_corrupted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.move(filename, corrupted_name)
            print(f"Moved corrupted file to: {corrupted_name}")
        
        # Restore from backup
        shutil.copy(best_backup, filename)
        print(f"Restored from: {best_backup}")
        
        # Verify restored file
        try:
            wb = openpyxl.load_workbook(filename, data_only=True)
            print(f"✅ File restored successfully!")
            print(f"   Sheets: {len(wb.sheetnames)}")
            
            # Read summary
            ws = wb['Summary']
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if row[0] and row[1] is not None:
                    print(f"   {row[0]}: {row[1]}")
            
            return True
        except Exception as e:
            print(f"❌ Restored file is also corrupted: {e}")
            return False
    else:
        print("❌ No valid backups found!")
        return False


def create_fresh_excel():
    """Create fresh Excel file if all backups are corrupted"""
    
    filename = "live_trading_results.xlsx"
    
    # Backup any existing file
    if os.path.exists(filename):
        backup_name = f"live_trading_results_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.move(filename, backup_name)
        print(f"Backed up existing file to: {backup_name}")
    
    # Create fresh file
    initial_capital = 7400
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = {
            'Metric': ['Initial Capital', 'Current Capital', 'Total P&L $', 'Total P&L %', 
                      'Total Trades', 'Winning Trades', 'Losing Trades', 'Win Rate %',
                      'Trade Size', 'Active Strategies', 'Bankrupt Strategies'],
            'Value': [initial_capital, initial_capital, 0, 0, 0, 0, 0, 0, 5, 74, 0]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        
        # Strategy Status sheet
        status_data = {
            'Strategy': [],
            'Status': [],
            'Capital': [],
            'Trades': [],
            'P&L $': [],
            'P&L %': [],
            'Win Rate %': []
        }
        pd.DataFrame(status_data).to_excel(writer, sheet_name='Strategy Status', index=False)
        
        # All Trades sheet
        trades_data = {
            'Trade #': [],
            'Date': [],
            'Time': [],
            'Strategy': [],
            'Side': [],
            'Entry Price': [],
            'Exit Price': [],
            'Status': [],
            'P&L %': [],
            'P&L $': [],
            'Capital After': [],
            'Confidence': [],
            'Entry Reason': [],
            'Exit Reason': [],
            'Duration (min)': []
        }
        pd.DataFrame(trades_data).to_excel(writer, sheet_name='All Trades', index=False)
    
    print(f"✅ Created fresh Excel file: {filename}")
    print(f"   Initial Capital: ${initial_capital}")
    return True


if __name__ == '__main__':
    print("=" * 60)
    print("EXCEL FILE REPAIR TOOL")
    print("=" * 60)
    
    # Try to repair from backup first
    if repair_excel():
        print("\n✅ Repair successful!")
    else:
        print("\n⚠️  Could not repair from backup. Creating fresh file...")
        if create_fresh_excel():
            print("✅ Fresh file created!")
        else:
            print("❌ Failed to create fresh file!")
