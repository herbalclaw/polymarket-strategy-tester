#!/usr/bin/env python3
"""
Hourly Progress Report Generator
Generates accurate P&L reports with correct percentage calculations
"""

import openpyxl
import os
from datetime import datetime

def generate_hourly_report():
    """Generate hourly progress report with correct calculations"""
    
    filename = "live_trading_results.xlsx"
    
    if not os.path.exists(filename):
        return "Error: Excel file not found"
    
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        
        # Read Summary
        summary = {}
        ws_summary = wb['Summary']
        for row in ws_summary.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                summary[row[0]] = row[1]
        
        # Read Strategy Status
        ws_status = wb['Strategy Status']
        strategies = []
        for row in ws_status.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                strategies.append({
                    'name': row[0],
                    'status': row[1],
                    'capital': row[2] if row[2] is not None else 100,
                    'trades': row[3] if row[3] is not None else 0,
                    'pnl': row[4] if row[4] is not None else 0,
                    'pnl_pct': row[5] if row[5] is not None else 0,
                    'win_rate': row[6] if row[6] is not None else 0
                })
        
        # Sort by P&L
        strategies.sort(key=lambda x: x['pnl'], reverse=True)
        
        # Calculate correct percentages
        # P&L % should be (P&L $ / Initial Capital) * 100
        initial_capital = summary.get('Initial Capital', 7400)
        total_pnl = summary.get('Total P&L $', 0)
        correct_total_pct = (total_pnl / initial_capital * 100) if initial_capital > 0 else 0
        
        # Generate report
        report = []
        report.append("🕐 Hourly Progress Report — " + datetime.now().strftime("%I:%M %p (Asia/Shanghai)"))
        report.append("")
        report.append("📊 Portfolio Summary:")
        report.append(f"  Initial Capital: ${initial_capital:,.2f}")
        report.append(f"  Current Capital: ${summary.get('Current Capital', initial_capital):,.2f}")
        report.append(f"  Total P&L: ${total_pnl:+.2f} ({correct_total_pct:+.2f}%)")
        report.append(f"  Total Trades: {int(summary.get('Total Trades', 0))}")
        report.append(f"  Win Rate: {summary.get('Win Rate %', 0):.1f}%")
        report.append("")
        
        # Top strategies
        report.append("🏆 Top Strategies:")
        for i, s in enumerate(strategies[:8], 1):
            if s['trades'] > 0:  # Only show strategies with trades
                # Correct P&L % calculation: (P&L $ / Strategy Initial Capital) * 100
                strat_initial = 100  # Each strategy starts with $100
                correct_pct = (s['pnl'] / strat_initial * 100) if strat_initial > 0 else 0
                emoji = "🟢" if s['pnl'] >= 0 else "🔴"
                report.append(f"  {i}. {s['name'][:25]:<25} {s['trades']:>3} trades  ${s['pnl']:>+6.2f} ({correct_pct:>+5.2f}%) {emoji}")
        
        report.append("")
        
        # Bottom strategies
        losers = [s for s in strategies if s['pnl'] < 0 and s['trades'] > 0]
        if losers:
            report.append("📉 Losing Strategies:")
            for s in losers[:5]:
                strat_initial = 100
                correct_pct = (s['pnl'] / strat_initial * 100) if strat_initial > 0 else 0
                report.append(f"  • {s['name'][:25]:<25} {s['trades']:>3} trades  ${s['pnl']:>+6.2f} ({correct_pct:>+5.2f}%)")
            report.append("")
        
        # Scalper status
        scalpers = [s for s in strategies if 'Scalper' in s['name']]
        active_scalpers = [s for s in scalpers if s['trades'] > 0]
        report.append("⚡ Scalper Strategies:")
        report.append(f"  Total: {len(scalpers)} | Active (with trades): {len(active_scalpers)}")
        for s in scalpers:
            if s['trades'] > 0:
                strat_initial = 100
                correct_pct = (s['pnl'] / strat_initial * 100) if strat_initial > 0 else 0
                report.append(f"  • {s['name'][:20]:<20} {s['trades']:>3} trades  ${s['pnl']:>+6.2f} ({correct_pct:>+5.2f}%)")
        report.append("")
        
        report.append("✅ Report generated with correct percentage calculations")
        report.append("   (P&L % = P&L $ / Strategy Initial Capital × 100)")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Error generating report: {type(e).__name__}: {e}"


if __name__ == '__main__':
    print(generate_hourly_report())
